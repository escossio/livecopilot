import argparse
import csv
import json
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from app.services.job_market import normalize_text


@dataclass(frozen=True)
class TermSpec:
    term: str
    category: str
    subcategory: str
    match_groups: Tuple[Tuple[str, ...], ...]
    aliases: Tuple[str, ...] = ()
    excludes: Tuple[str, ...] = ()
    secondary_category: Optional[str] = None
    secondary_subcategory: Optional[str] = None
    generic: bool = False


GENERIC_STOPWORDS = {
    "pleno",
    "senior",
    "sênior",
    "júnior",
    "junior",
    "jr",
    "sr",
    "analista",
    "assistente",
    "vaga",
    "vagas",
    "especialista",
    "coordenador",
    "coordenadora",
    "gestor",
    "gerente",
    "engenheiro",
    "engenheira",
    "programador",
    "programadora",
    "desenvolvedor",
    "desenvolvedora",
    "developer",
}

BACKEND_MARKERS = ("backend", "back end", "back-end", "api")
FRONTEND_MARKERS = ("frontend", "front end", "front-end")
PRODUCT_GENERIC_EXCLUDES = (
    "product owner",
    "product manager",
    "product designer",
    "product design",
    "product ops",
    "product operations",
)
CLOUD_PROVIDER_EXCLUDES = ("aws", "amazon web services", "azure", "gcp", "google cloud", "google cloud platform")
DATA_SPECIFIC_EXCLUDES = (
    "data engineer",
    "engenheiro de dados",
    "engenharia de dados",
    "data science",
    "data scientist",
    "cientista de dados",
    "analytics",
    "business intelligence",
    "power bi",
    "tableau",
    "machine learning",
    "inteligencia artificial",
    "artificial intelligence",
    "ml engineer",
    "genai",
    "llm",
)
BACKEND_LANGUAGE_EXCLUDES = (
    "java",
    "spring",
    ".net",
    "dotnet",
    "c#",
    "c sharp",
    "python",
    "django",
    "fastapi",
    "flask",
    "php",
    "laravel",
    "symfony",
    "node.js",
    "node js",
    "nodejs",
    "nestjs",
    "express",
)

TERM_SPECS: List[TermSpec] = [
    TermSpec(
        "data engineering",
        "dados",
        "data engineering",
        (("data engineering", "data engineer", "engenharia de dados", "engenheiro de dados", "etl", "elt"),),
        aliases=("engenharia de dados", "data engineer", "engenheiro de dados", "etl", "elt"),
    ),
    TermSpec(
        "data science",
        "dados",
        "data science",
        (("data science", "data scientist", "cientista de dados"),),
        aliases=("data science", "data scientist", "cientista de dados"),
    ),
    TermSpec(
        "analytics / bi",
        "dados",
        "analytics / bi",
        (("analytics", "business intelligence", "power bi", "tableau", "analista de dados", "analista bi", "bi"),),
        aliases=("analytics", "business intelligence", "power bi", "tableau", "bi", "analista de dados"),
    ),
    TermSpec(
        "ia / ml",
        "dados",
        "ia / ml",
        (("inteligencia artificial", "inteligência artificial", "artificial intelligence", "machine learning", "aprendizado de maquina", "aprendizado de máquina", "ml engineer", "genai", "llm", "nlp", "computer vision"),),
        aliases=("ia", "ai", "machine learning", "ml", "genai", "llm", "inteligencia artificial"),
    ),
    TermSpec(
        "dados",
        "dados",
        "geral",
        (("dados",),),
        excludes=DATA_SPECIFIC_EXCLUDES,
        aliases=("dados", "data"),
        generic=True,
    ),
    TermSpec("full stack", "stack", "full stack", (("full stack", "fullstack"),), aliases=("full stack", "fullstack")),
    TermSpec(
        "backend java",
        "backend",
        "java",
        (BACKEND_MARKERS, ("java", "spring")),
        aliases=("backend java", "java backend", "spring backend", "api java"),
    ),
    TermSpec(
        "backend .net",
        "backend",
        ".net",
        (BACKEND_MARKERS, (".net", "dotnet", "c#", "c sharp")),
        aliases=("backend .net", ".net backend", "dotnet backend", "c# backend"),
    ),
    TermSpec(
        "backend python",
        "backend",
        "python",
        (BACKEND_MARKERS, ("python", "django", "fastapi", "flask")),
        aliases=("backend python", "python backend", "django", "fastapi", "flask"),
    ),
    TermSpec(
        "backend php",
        "backend",
        "php",
        (BACKEND_MARKERS, ("php", "laravel", "symfony")),
        aliases=("backend php", "php backend", "laravel", "symfony"),
    ),
    TermSpec(
        "backend node",
        "backend",
        "node",
        (BACKEND_MARKERS, ("node.js", "node js", "nodejs", "nestjs", "express")),
        aliases=("backend node", "node backend", "node.js", "nodejs", "nestjs", "express"),
    ),
    TermSpec(
        "backend",
        "backend",
        "geral",
        (BACKEND_MARKERS,),
        excludes=BACKEND_LANGUAGE_EXCLUDES,
        aliases=("backend", "back end", "api"),
        generic=True,
    ),
    TermSpec("frontend", "stack", "frontend", (FRONTEND_MARKERS,), aliases=("frontend", "front end", "front-end")),
    TermSpec(
        "aws",
        "cloud",
        "aws",
        (("aws", "amazon web services"),),
        aliases=("aws", "amazon web services"),
    ),
    TermSpec("azure", "cloud", "azure", (("azure",),), aliases=("azure",)),
    TermSpec(
        "gcp",
        "cloud",
        "gcp",
        (("gcp", "google cloud", "google cloud platform"),),
        aliases=("gcp", "google cloud", "google cloud platform"),
    ),
    TermSpec(
        "cloud",
        "cloud",
        "genérico",
        (("cloud",),),
        excludes=CLOUD_PROVIDER_EXCLUDES,
        aliases=("cloud",),
        generic=True,
    ),
    TermSpec("devops", "devops", "devops", (("devops", "dev ops", "sre", "site reliability"),), aliases=("devops", "sre")),
    TermSpec("react", "frontend", "react", (("react", "reactjs", "react.js"),), aliases=("react", "reactjs")),
    TermSpec("angular", "frontend", "angular", (("angular", "angularjs", "angular.js"),), aliases=("angular", "angularjs")),
    TermSpec(
        "product owner",
        "produto",
        "product owner",
        (("product owner",),),
        aliases=("product owner", "po"),
    ),
    TermSpec(
        "product manager",
        "produto",
        "product manager",
        (("product manager", "gerente de produto", "head de produto"),),
        aliases=("product manager", "gerente de produto", "pm"),
    ),
    TermSpec(
        "product design",
        "produto",
        "product design",
        (("product designer", "product design", "designer de produto"),),
        aliases=("product designer", "product design", "designer de produto"),
        secondary_category="design",
        secondary_subcategory="product design",
    ),
    TermSpec(
        "product ops",
        "produto",
        "product ops",
        (("product ops", "product operations", "operacoes de produto", "operações de produto"),),
        aliases=("product ops", "product operations", "operações de produto"),
    ),
    TermSpec(
        "produto",
        "produto",
        "geral",
        (("product", "produto"),),
        excludes=PRODUCT_GENERIC_EXCLUDES,
        aliases=("product", "produto"),
        generic=True,
    ),
    TermSpec(
        "ux",
        "design",
        "ux",
        (("ux researcher", "ux research", "ux designer", "ux ui", "ui ux", "user experience"),),
        aliases=("ux", "ux designer", "ux research", "user experience"),
    ),
    TermSpec(
        "ui",
        "design",
        "ui",
        (("ui designer", "ux ui", "ui ux", "user interface"),),
        aliases=("ui", "ui designer", "user interface"),
    ),
    TermSpec(
        "motion / visual design",
        "design",
        "motion / visual design",
        (("motion designer", "visual designer", "graphic designer", "designer grafico", "designer gráfico"),),
        aliases=("motion designer", "visual designer", "graphic designer", "designer gráfico"),
    ),
    TermSpec("sql", "banco", "sql", (("sql", "mysql", "postgres", "postgresql", "mssql"),), aliases=("sql", "postgres", "mysql")),
    TermSpec("oracle", "banco", "oracle", (("oracle",),), aliases=("oracle",)),
    TermSpec("dba", "banco", "dba", (("dba", "database administrator"),), aliases=("dba", "database administrator")),
    TermSpec("segurança", "segurança", "segurança", (("seguranca", "segurança", "security", "infosec"),), aliases=("segurança", "security", "infosec")),
    TermSpec("infraestrutura", "infraestrutura", "infraestrutura", (("infraestrutura", "infra", "infra-estrutura"),), aliases=("infraestrutura", "infra")),
    TermSpec("testes", "qa", "qa", (("testes", "tester", "qa", "quality assurance", "test automation"),), aliases=("qa", "testes", "quality assurance")),
    TermSpec("suporte", "infraestrutura", "suporte", (("suporte", "service desk", "field service", "help desk"),), aliases=("suporte", "service desk", "field service")),
    TermSpec("sap", "erp", "sap", (("sap",),), aliases=("sap",)),
    TermSpec("datasul", "erp", "datasul", (("datasul",),), aliases=("datasul",)),
    TermSpec("salesforce", "crm", "salesforce", (("salesforce",),), aliases=("salesforce",)),
    TermSpec("pmo", "gestão", "pmo", (("pmo", "project management office"),), aliases=("pmo", "project management office")),
    TermSpec("vtex", "ecommerce", "vtex", (("vtex",),), aliases=("vtex",)),
    TermSpec("php", "linguagem", "php", (("php",),), aliases=("php",), generic=True),
    TermSpec("java", "linguagem", "java", (("java",),), aliases=("java",), generic=True),
    TermSpec(".net", "linguagem", ".net", ((".net", "dotnet", "c#", "c sharp"),), aliases=(".net", "dotnet", "c#"), generic=True),
    TermSpec("python", "linguagem", "python", (("python",),), aliases=("python",), generic=True),
    TermSpec("node.js", "linguagem", "node.js", (("node.js", "node js", "nodejs"),), aliases=("node.js", "nodejs"), generic=True),
    TermSpec("delphi", "linguagem", "delphi", (("delphi",),), aliases=("delphi",), generic=True),
]


def _load_xlsx_titles(path: Path, sheet_name: str) -> List[str]:
    try:
        import openpyxl
    except Exception:
        raise RuntimeError("openpyxl não instalado")

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    title_idx = _detect_title_index(headers)
    titles = []
    for row in rows[1:]:
        if title_idx is None:
            candidates = [cell for cell in row if isinstance(cell, str)]
            if candidates:
                titles.append(candidates[0])
            continue
        cell = row[title_idx] if title_idx < len(row) else None
        if isinstance(cell, str) and cell.strip():
            titles.append(cell.strip())
    return titles


def _load_csv_titles(path: Path) -> List[str]:
    titles: List[str] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames:
            headers = [field.strip().lower() for field in reader.fieldnames]
            title_key = _detect_title_key(headers)
            for row in reader:
                if title_key and row.get(title_key):
                    titles.append(str(row[title_key]).strip())
                else:
                    for value in row.values():
                        if isinstance(value, str) and value.strip():
                            titles.append(value.strip())
                            break
        else:
            handle.seek(0)
            raw_reader = csv.reader(handle)
            for row in raw_reader:
                if row:
                    titles.append(row[0].strip())
    return titles


def _detect_title_index(headers: List[str]) -> Optional[int]:
    for idx, header in enumerate(headers):
        if _is_title_header(header):
            return idx
    return None


def _detect_title_key(headers: List[str]) -> Optional[str]:
    for header in headers:
        if _is_title_header(header):
            return header
    return None


def _is_title_header(header: str) -> bool:
    header = header.strip().lower()
    return any(token in header for token in ["titulo", "título", "title", "vaga", "position"])


def _contains_phrase(normalized_text: str, phrase: str) -> bool:
    return f" {normalize_text(phrase)} " in normalized_text


def _matches_spec(normalized_title: str, spec: TermSpec) -> bool:
    if spec.excludes and any(_contains_phrase(normalized_title, phrase) for phrase in spec.excludes):
        return False
    for group in spec.match_groups:
        if not any(_contains_phrase(normalized_title, phrase) for phrase in group):
            return False
    return True


def _extract_terms(title: str) -> List[str]:
    normalized = f" {normalize_text(title)} "
    tokens = [token for token in normalized.split() if token not in GENERIC_STOPWORDS]
    filtered = f" {' '.join(tokens)} "
    matches = []
    for spec in TERM_SPECS:
        if _matches_spec(filtered, spec):
            matches.append(spec.term)
    return matches


def build_market_seed(titles: Iterable[str], max_examples: int = 5) -> List[Dict[str, object]]:
    counts: Dict[str, int] = defaultdict(int)
    examples: Dict[str, List[str]] = defaultdict(list)
    for title in titles:
        if not isinstance(title, str):
            continue
        title_clean = title.strip()
        if not title_clean:
            continue
        terms = _extract_terms(title_clean)
        for term in terms:
            counts[term] += 1
            if title_clean not in examples[term] and len(examples[term]) < max_examples:
                examples[term].append(title_clean)
    results = []
    for spec in TERM_SPECS:
        if counts.get(spec.term):
            results.append(
                {
                    "termo": spec.term,
                    "term": spec.term,
                    "frequencia": counts[spec.term],
                    "categoria": spec.category,
                    "subcategoria": spec.subcategory,
                    "frequency": counts[spec.term],
                    "category": spec.category,
                    "subcategory": spec.subcategory,
                    "aliases": list(dict.fromkeys((spec.term, *spec.aliases))),
                    "generic": spec.generic,
                    "examples_de_titulos": examples[spec.term],
                }
            )
            if spec.secondary_category:
                results[-1]["categoria_secundaria"] = spec.secondary_category
                results[-1]["secondary_category"] = spec.secondary_category
            if spec.secondary_subcategory:
                results[-1]["subcategoria_secundaria"] = spec.secondary_subcategory
                results[-1]["secondary_subcategory"] = spec.secondary_subcategory
    results.sort(key=lambda item: (-item["frequency"], item["termo"]))
    return results


def load_titles(xlsx_path: Path, csv_path: Path, sheet_name: str) -> Tuple[List[str], str]:
    if xlsx_path.exists():
        try:
            return _load_xlsx_titles(xlsx_path, sheet_name), "xlsx"
        except Exception:
            pass
    if csv_path.exists():
        return _load_csv_titles(csv_path), "csv"
    return [], "none"


def write_seed_json(output_path: Path, payload: List[Dict[str, object]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Ingestao inicial de vagas para memoria de mercado.")
    parser.add_argument("--xlsx", default="data/vagas.xlsx", help="Caminho para o XLSX de vagas.")
    parser.add_argument("--csv", default="data/vagas.csv", help="Caminho para o CSV de vagas.")
    parser.add_argument("--sheet", default="Todas as Vagas", help="Nome da aba preferencial no XLSX.")
    parser.add_argument("--output", default="data/job_market_seed.json", help="Arquivo JSON de saida.")
    parser.add_argument("--max-examples", type=int, default=5, help="Maximo de exemplos por termo.")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    csv_path = Path(args.csv)
    titles, source = load_titles(xlsx_path, csv_path, args.sheet)
    if not titles:
        print("Nenhum titulo encontrado. Verifique o XLSX/CSV informado.")
        return 1

    payload = build_market_seed(titles, max_examples=args.max_examples)
    write_seed_json(Path(args.output), payload)

    top = payload[:10]
    print(f"Fonte usada: {source}")
    print(f"Titulos processados: {len(titles)}")
    print("Top termos:")
    for item in top:
        print(f"- {item['termo']}: {item['frequencia']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
